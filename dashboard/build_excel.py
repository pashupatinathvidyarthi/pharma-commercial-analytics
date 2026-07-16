"""
Builds the Excel deliverable (Commercial_Analytics_Report.xlsx) using openpyxl,
with live SUMIFS/COUNTIFS formulas referencing a raw-data tab — matching the
Excel skill the JD explicitly asks for, alongside SQL and Python.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA = Path(__file__).resolve().parent.parent / "data"
OUT_PATH = Path(__file__).resolve().parent.parent / "Commercial_Analytics_Report.xlsx"

raw = pd.read_csv(DATA / "model_outputs" / "excel_raw_data.csv")

wb = Workbook()

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BODY_FONT = Font(name="Arial")
TITLE_FONT = Font(name="Arial", bold=True, size=14)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

def autofit(ws, ncols, width=16):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width

# ------------------------------------------------------------------
# Sheet 1: Raw_Data
# ------------------------------------------------------------------
ws_raw = wb.active
ws_raw.title = "Raw_Data"
cols = list(raw.columns)
ws_raw.append(cols)
style_header(ws_raw, 1, len(cols))
for r in raw.itertuples(index=False):
    ws_raw.append(list(r))
for row in ws_raw.iter_rows(min_row=2, max_row=ws_raw.max_row):
    for cell in row:
        cell.font = BODY_FONT
autofit(ws_raw, len(cols), 14)
N = ws_raw.max_row  # last data row

col_idx = {name: i + 1 for i, name in enumerate(cols)}
def L(name):
    return get_column_letter(col_idx[name])

# ------------------------------------------------------------------
# Sheet 2: Segmentation_Summary (SUMIFS / COUNTIFS / AVERAGEIFS)
# ------------------------------------------------------------------
ws_seg = wb.create_sheet("Segmentation_Summary")
ws_seg["A1"] = "HCP Segmentation Summary"
ws_seg["A1"].font = TITLE_FONT
headers = ["Segment", "HCP Count", "Avg Potential", "Avg Receptivity", "Total Rx Volume", "Avg Calls/HCP"]
ws_seg.append([])
ws_seg.append(headers)
style_header(ws_seg, 3, len(headers))

segments = ["Key Target", "Growth", "Maintain", "Low Priority"]
raw_rng = f"Raw_Data!${L('segment')}$2:${L('segment')}${N}"
for i, seg in enumerate(segments, start=4):
    ws_seg.cell(row=i, column=1, value=seg)
    ws_seg.cell(row=i, column=2, value=f'=COUNTIF({raw_rng},A{i})')
    ws_seg.cell(row=i, column=3, value=f'=AVERAGEIFS(Raw_Data!${L("potential_score")}$2:${L("potential_score")}${N},{raw_rng},A{i})')
    ws_seg.cell(row=i, column=4, value=f'=AVERAGEIFS(Raw_Data!${L("receptivity_score")}$2:${L("receptivity_score")}${N},{raw_rng},A{i})')
    ws_seg.cell(row=i, column=5, value=f'=SUMIFS(Raw_Data!${L("total_rx")}$2:${L("total_rx")}${N},{raw_rng},A{i})')
    ws_seg.cell(row=i, column=6, value=f'=AVERAGEIFS(Raw_Data!${L("total_calls")}$2:${L("total_calls")}${N},{raw_rng},A{i})')
    for c in range(1, 7):
        ws_seg.cell(row=i, column=c).font = BODY_FONT
        if c in (3, 4, 6):
            ws_seg.cell(row=i, column=c).number_format = "0.0"
        if c == 5:
            ws_seg.cell(row=i, column=c).number_format = "#,##0"

# Total row
tot_row = 4 + len(segments)
ws_seg.cell(row=tot_row, column=1, value="Total / Overall").font = Font(name="Arial", bold=True)
ws_seg.cell(row=tot_row, column=2, value=f"=SUM(B4:B{tot_row-1})")
ws_seg.cell(row=tot_row, column=5, value=f"=SUM(E4:E{tot_row-1})").number_format = "#,##0"
autofit(ws_seg, 6, 18)

# ------------------------------------------------------------------
# Sheet 3: Call_Plan_Compliance
# ------------------------------------------------------------------
ws_call = wb.create_sheet("Call_Plan_Compliance")
ws_call["A1"] = "Call Plan Compliance by Segment"
ws_call["A1"].font = TITLE_FONT
headers2 = ["Segment", "Target Calls/Month", "Actual Calls/Month", "Compliance %"]
ws_call.append([])
ws_call.append(headers2)
style_header(ws_call, 3, len(headers2))

targets = {"Key Target": 2.0, "Growth": 1.5, "Maintain": 1.0, "Low Priority": 0.4}
for i, seg in enumerate(segments, start=4):
    ws_call.cell(row=i, column=1, value=seg)
    ws_call.cell(row=i, column=2, value=targets[seg])
    ws_call.cell(row=i, column=3, value=f'=Segmentation_Summary!F{i}/12').number_format = "0.00"
    ws_call.cell(row=i, column=4, value=f'=C{i}/B{i}').number_format = "0.0%"
    for c in range(1, 5):
        ws_call.cell(row=i, column=c).font = BODY_FONT
autofit(ws_call, 4, 20)

# ------------------------------------------------------------------
# Sheet 4: Marketing_Mix (Specialty x Region pivot via SUMIFS)
# ------------------------------------------------------------------
ws_mm = wb.create_sheet("Marketing_Mix")
ws_mm["A1"] = "Rx per Call — Specialty x Region"
ws_mm["A1"].font = TITLE_FONT
specialties = sorted(raw["specialty"].unique())
regions = sorted(raw["region"].unique())
ws_mm.append([])
ws_mm.append(["Specialty"] + regions)
style_header(ws_mm, 3, len(regions) + 1)

spec_rng = f"Raw_Data!${L('specialty')}$2:${L('specialty')}${N}"
reg_rng = f"Raw_Data!${L('region')}$2:${L('region')}${N}"
rx_rng = f"Raw_Data!${L('total_rx')}$2:${L('total_rx')}${N}"
calls_rng = f"Raw_Data!${L('total_calls')}$2:${L('total_calls')}${N}"

for r, spec in enumerate(specialties, start=4):
    ws_mm.cell(row=r, column=1, value=spec).font = BODY_FONT
    for c, reg in enumerate(regions, start=2):
        col_letter = get_column_letter(c)
        formula = (
            f'=IFERROR(SUMIFS({rx_rng},{spec_rng},$A{r},{reg_rng},{col_letter}$3)'
            f'/SUMIFS({calls_rng},{spec_rng},$A{r},{reg_rng},{col_letter}$3),0)'
        )
        cell = ws_mm.cell(row=r, column=c, value=formula)
        cell.number_format = "0.0"
        cell.font = BODY_FONT
autofit(ws_mm, len(regions) + 1, 14)

# ------------------------------------------------------------------
# Sheet 5: ICM_Assumptions (documented, editable inputs — yellow fill)
# ------------------------------------------------------------------
ws_icm = wb.create_sheet("ICM_Assumptions")
ws_icm["A1"] = "Incentive Compensation — Payout Curve Assumptions"
ws_icm["A1"].font = TITLE_FONT
YELLOW = PatternFill("solid", fgColor="FFFF00")
rows = [
    ("Threshold (min payout starts)", 0.80, "Below this attainment: $0 payout"),
    ("Target attainment", 1.00, "100% attainment = Full Payout"),
    ("Accelerator threshold", 1.10, "At/above this attainment = Accelerator tier"),
    ("Base Quarterly Incentive ($)", 5000, "Illustrative target incentive per rep per quarter"),
    ("Partial Payout %", 0.50, "Payout multiple for Threshold <= attainment < 100%"),
    ("Full Payout %", 1.00, "Payout multiple for 100% <= attainment < Accelerator"),
    ("Accelerator Payout %", 1.50, "Payout multiple at/above Accelerator threshold"),
]
ws_icm.append([])
ws_icm.append(["Assumption", "Value", "Notes"])
style_header(ws_icm, 3, 3)
for i, (label, val, note) in enumerate(rows, start=4):
    ws_icm.cell(row=i, column=1, value=label).font = BODY_FONT
    vcell = ws_icm.cell(row=i, column=2, value=val)
    vcell.font = Font(name="Arial", color="0000FF")
    vcell.fill = YELLOW
    if val < 10:
        vcell.number_format = "0%" if val <= 1 else "0.00"
    ws_icm.cell(row=i, column=3, value=note).font = Font(name="Arial", italic=True)
autofit(ws_icm, 3, 30)
ws_icm.column_dimensions["C"].width = 45

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")

